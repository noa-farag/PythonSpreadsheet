import json
from table_calculator import TableCalculator
import pandas as pd
from cell import Cell
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from typing import List


table = TableCalculator
Location = tuple[int, int]


class File:
    """
    class that handles the file operations
    """

    def save_to_json(self, data: dict, file_path: str):
        """
        save the data to a json file
        in the format of [title, {(0, 1): [str(seen_value), str(formula), str(color)}]]
        :param data: dictionary to save
        :param file_path: where to save the file
        """
        # make sure the file has the right format
        try:
            if file_path.endswith('.json'):
                file_path = file_path
            else:
                file_path = file_path + '.json'
            # save the data to the file
            with open(file_path, 'w') as file:
                json.dump(data, file, indent=4)
        except Exception:
            raise Exception

    def load_json_data(self, json_filepath) -> dict:
        """
        load the data from given json path
        works only with format of [title, {(0, 1): str(formula), str(color)}] format
        :param json_filepath: the path to the json file
        :return: the data from the file as dictionary
        """
        try:
            with open(json_filepath, 'r') as file:
                data: dict = json.load(file)
            return data
        except Exception:
            raise OSError

    def export_data_to_excel(self, data: List[List[Cell]], data_frame: pd.DataFrame, file_path: str):
        """
        save the dataframe to excel
        :param data: the data to save
        :param data_frame: the dataframe to save
        :param file_path: where to save the file
        """
        try:
            file_path = file_path + '.xlsx'
            # convert number strings to int or float
            data_frame = data_frame.applymap(lambda x: pd.to_numeric(x, errors='ignore'))  # type: ignore
            data_frame.to_excel(file_path, index=False, header=False)
            wb = load_workbook(file_path)
            ws = wb.active
            # change the color of the cells in the Excel file
            for i in range(len(data)):
                for j in range(len(data[0])):
                    cell: Cell = data[i][j]
                    color: str = cell.get_color()[1:]
                    cell_address: str = ws.cell(i + 1, j + 1).coordinate  # type: ignore
                    if color != 'ffffff':
                        fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        ws[cell_address].fill = fill
            wb.save(file_path)
        except Exception:
            raise Exception
